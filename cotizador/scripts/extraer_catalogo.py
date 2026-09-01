#!/usr/bin/env python3
"""
Convierte el LISTADO DE PRECIOS en Excel al catálogo JSON que consume el
cotizador.

El Excel es un documento de trabajo, no una base de datos: el nombre de la
referencia sólo aparece en la primera fila de cada bloque, las filas siguientes
son escalones de cantidad del mismo producto, y en medio hay notas, filas de
clisé y celdas de cantidad escritas a mano ("DESDE 1000"). Todo eso se
normaliza aquí, una sola vez, para que la aplicación no tenga que saber nada
del Excel.

Uso:
    python3 scripts/extraer_catalogo.py <archivo.xlsx> [-o src/data/catalogo.json]
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import unicodedata
import warnings
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

import openpyxl

warnings.filterwarnings("ignore", message=".*Data Validation.*")

# --- Hojas del libro -------------------------------------------------------

HOJA_PRECIOS = "LISTADO PRECIOS 2026"
HOJA_ETIQUETAS = "precios etiquetas"
HOJA_SATELITALES = "SATELITALES"

# Nombres comerciales, fuera de este script porque cambian con el catálogo
# comercial y no con los precios. Ver la cabecera del propio archivo.
NOMBRES_COMERCIALES = Path(__file__).resolve().parent.parent / "src/datos/nombres-comerciales.json"

IVA = 0.19

# Filas cuyo texto en la columna de referencia no es un producto sino un
# encabezado repetido o una nota suelta.
RUIDO_REFERENCIA = re.compile(
    r"^(referencia|valor venta|valor compra|cantidad|total|nota|observaci)", re.I
)

# Filas que son un servicio de personalización y no un escalón de precio del
# producto que las precede.
PATRON_CLISE = re.compile(r"clis[eé]", re.I)

# Erratas del Excel que rompen la búsqueda por nombre. Se corrigen al leer y no
# en el archivo original, que lo mantiene el área comercial.
ERRATAS: list[tuple[str, str]] = [
    (r"\bGAYA\b", "GUAYA"),
    (r"\bREFE\b", "REF."),
    (r"\bDELLO\b", "SELLO"),
    (r"\bSIJETADORES\b", "SUJETADORES"),
]

# La columna de cantidad del proveedor es polimórfica: a veces es un número,
# a veces el empaque ("CAJA X 25 UNIDADES") y a veces el código del proveedor
# ("REF. 685"). El empaque sí interesa: hay producto que sólo se vende por caja.
PATRON_EMPAQUE = re.compile(
    r"\b(caja|paquete|bolsa|rollo)s?\s*x\s*([\d.,]+)", re.I
)


# --- Utilidades ------------------------------------------------------------


def texto(valor) -> str:
    """Normaliza una celda a texto limpio y colapsa espacios internos."""
    if valor is None:
        return ""
    if isinstance(valor, datetime):
        return valor.date().isoformat()
    return re.sub(r"\s+", " ", str(valor)).strip()


def numero(valor) -> float | None:
    """Lee una celda numérica tolerando texto como '$ 1.200' o 'DESDE 1000'."""
    if valor is None:
        return None
    if isinstance(valor, (int, float)):
        return float(valor)
    crudo = str(valor).strip()
    if not crudo:
        return None
    # Formato colombiano: punto de miles, coma decimal.
    limpio = re.sub(r"[^\d,.\-]", "", crudo)
    if not re.search(r"\d", limpio):
        return None
    if "," in limpio and "." in limpio:
        limpio = limpio.replace(".", "").replace(",", ".")
    elif "," in limpio:
        limpio = limpio.replace(",", ".")
    elif limpio.count(".") > 1:
        limpio = limpio.replace(".", "")
    try:
        return float(limpio)
    except ValueError:
        return None


def entero(valor) -> int | None:
    n = numero(valor)
    return int(round(n)) if n is not None else None


def cargar_nombres_comerciales() -> dict[str, dict]:
    """El nombre con el que el cliente lee cada referencia en su cotización.

    El listado de precios dice "PRECINTO ANCLA CAJAS"; la dueña vende
    "Precinto Ancla para Cajas de Seguridad Plásticas". Los dos nombres son
    suyos —el primero es como compra, el segundo es como vende— y el catálogo
    guarda los dos: el comercial en `nombre`, que es lo que sale en el PDF, y
    la referencia en `referencia`, que es como el asesor busca.

    Aplicarlo AQUÍ es lo que hace que el trabajo sobreviva al listado del año
    que viene. Sin esto, cada regeneración devolvía las 107 referencias a
    mayúsculas y había que rehacerlo a mano.
    """
    if not NOMBRES_COMERCIALES.exists():
        print(f"AVISO: no está {NOMBRES_COMERCIALES.name} — el catálogo saldrá con los nombres del listado de precios.")
        return {}
    datos = json.loads(NOMBRES_COMERCIALES.read_text(encoding="utf-8"))
    return datos.get("productos", {})


def corregir(nombre: str) -> str:
    """Aplica el diccionario de erratas al nombre de una referencia."""
    for patron, reemplazo in ERRATAS:
        nombre = re.sub(patron, reemplazo, nombre, flags=re.I)
    return re.sub(r"\s+", " ", nombre).strip()


def empaque_de(valor) -> str | None:
    """'CAJA X 25 UNIDADES' → 'Caja x 25'. Devuelve None si no es un empaque."""
    coincidencia = PATRON_EMPAQUE.search(texto(valor))
    if not coincidencia:
        return None
    unidad = coincidencia.group(1).capitalize()
    cantidad = coincidencia.group(2).replace(".", "").replace(",", "")
    return f"{unidad} x {cantidad}"


def slug(nombre: str) -> str:
    """Identificador estable y legible a partir del nombre de la referencia."""
    base = unicodedata.normalize("NFKD", nombre).encode("ascii", "ignore").decode()
    base = re.sub(r"[^a-zA-Z0-9]+", "-", base).strip("-").lower()
    return re.sub(r"-{2,}", "-", base)


def es_si(valor) -> bool | None:
    t = texto(valor).upper()
    if t.startswith("SI") or t.startswith("SÍ"):
        return True
    if t.startswith("NO"):
        return False
    return None


# --- Categorización --------------------------------------------------------

# El Excel no tiene columna de categoría: se deriva del nombre de la
# referencia. Gana la primera regla que coincida, así que el orden es la
# regla de desempate — "BOLSA INFLABLE" es air bag y no bolsa de seguridad,
# "PRECINTO ANCLA CAJAS" es precinto y no caja.
REGLAS_CATEGORIA: list[tuple[str, str]] = [
    (r"clis[eé]|troquel|grabado|marcaci", "Servicios de personalización"),
    (r"gps|satelital|jt7\d", "Rastreo satelital"),
    (r"absorbente|humedad|s[ií]lica", "Control de humedad"),
    (r"tula", "Tulas de seguridad"),
    (r"etiqueta|void|holograma|adhesiv|sticker", "Etiquetas de seguridad"),
    (r"(correa|amarre) de identificaci", "Identificación"),
    (r"pin(es)?\b", "Accesorios"),
    (r"protector", "Accesorios"),
    (r"kit", "Kits de seguridad"),
    (r"chapa|candado|cerradura", "Chapas y candados"),
    (r"\bg[au]aya\b|botella", "Precintos de guaya"),
    (r"met[aá]lico|nylon|cable", "Precintos metálicos y de cable"),
    (r"precinto", "Precintos plásticos"),
    (r"infla(ble|dora|tor)|air ?bag|pistola|batex", "Air bags e infladores"),
    (r"cinta", "Cintas de seguridad"),
    (r"bolsa|courier|sobre", "Bolsas y sobres de seguridad"),
    (r"zuncho|fleje|esquinero|amarre", "Sujeción de carga"),
    (r"caja|contenedor", "Cajas y contenedores"),
]

CATEGORIA_POR_DEFECTO = "Otros productos"

# Orden comercial con el que se agrupan las categorías en la interfaz y en el
# PDF. Lo que no esté aquí queda al final, alfabético.
ORDEN_CATEGORIAS: list[str] = [
    "Precintos plásticos",
    "Precintos de guaya",
    "Precintos metálicos y de cable",
    "Chapas y candados",
    "Etiquetas de seguridad",
    "Cintas de seguridad",
    "Bolsas y sobres de seguridad",
    "Tulas de seguridad",
    "Air bags e infladores",
    "Sujeción de carga",
    "Cajas y contenedores",
    "Control de humedad",
    "Identificación",
    "Kits de seguridad",
    "Accesorios",
    "Rastreo satelital",
    "Servicios de personalización",
    CATEGORIA_POR_DEFECTO,
]


def categoria_de(nombre: str) -> str:
    for patron, etiqueta in REGLAS_CATEGORIA:
        if re.search(patron, nombre, re.I):
            return etiqueta
    return CATEGORIA_POR_DEFECTO


def peso_categoria(categoria: str) -> tuple[int, str]:
    if categoria in ORDEN_CATEGORIAS:
        return (ORDEN_CATEGORIAS.index(categoria), "")
    return (len(ORDEN_CATEGORIAS), categoria)


# --- Modelo intermedio -----------------------------------------------------


@dataclass
class Escalon:
    """Un renglón de precio: a partir de `cantidad` unidades vale `unitario`."""

    cantidad: int
    unitario: float
    logo: bool | None = None
    medida: str | None = None
    nota: str | None = None

    def clave(self) -> tuple:
        return (self.cantidad, bool(self.logo), self.medida or "")


@dataclass
class Producto:
    id: str
    nombre: str
    categoria: str
    escalones: list[Escalon] = field(default_factory=list)
    costos: list[dict] = field(default_factory=list)
    proveedor: str | None = None
    notas: list[str] = field(default_factory=list)
    # Medida → recargo por centímetro adicional sobre esa medida base. El
    # recargo no es único por producto: baja a medida que la etiqueta es más
    # larga, así que va por medida y no en el producto.
    medidas: dict[str, float | None] = field(default_factory=dict)
    empaque: str | None = None
    origen: str = HOJA_PRECIOS


class Recolector:
    """Acumula productos evitando duplicados de nombre entre hojas."""

    def __init__(self) -> None:
        self.productos: dict[str, Producto] = {}
        self.incidencias: list[dict] = []

    def obtener(self, nombre_crudo: str, origen: str) -> Producto:
        nombre = corregir(nombre_crudo)
        clave = slug(nombre)
        if clave not in self.productos:
            self.productos[clave] = Producto(
                id=clave,
                nombre=nombre,
                categoria=categoria_de(nombre),
                origen=origen,
            )
        return self.productos[clave]

    def anotar(self, tipo: str, detalle: str, **extra) -> None:
        self.incidencias.append({"tipo": tipo, "detalle": detalle, **extra})


# --- Lectura de la hoja principal -----------------------------------------


def leer_listado_precios(hoja, rec: Recolector) -> None:
    """
    Columnas: A referencia · B cantidad · C logo · D unitario · E subtotal ·
    F iva · G total · H observaciones ‖ K referencia proveedor · L cantidad ·
    M costo unitario · N/O notas sueltas.
    """
    actual: Producto | None = None
    proveedor_actual: str | None = None
    # La columna LOGO se diligencia una sola vez, en la fila donde arranca el
    # sub-bloque, y rige para las filas siguientes. Varias referencias
    # aparecen dos veces seguidas —una tanda CON logo y otra SIN logo— y sin
    # arrastrar este valor las dos tandas se confunden en un mismo escalón.
    logo_actual: bool | None = None

    for fila in hoja.iter_rows(min_row=3, values_only=True):
        celdas = list(fila) + [None] * (15 - len(fila))
        referencia = texto(celdas[0])
        cantidad_cruda = celdas[1]
        cantidad_texto = texto(cantidad_cruda)
        unitario = numero(celdas[3])
        observacion = texto(celdas[7])
        ref_proveedor = texto(celdas[10])
        costo = numero(celdas[12])

        if referencia and not RUIDO_REFERENCIA.match(referencia):
            actual = rec.obtener(referencia, HOJA_PRECIOS)
            proveedor_actual = None
            logo_actual = None

        logo_fila = es_si(celdas[2])
        if logo_fila is not None:
            logo_actual = logo_fila

        if ref_proveedor and actual and not RUIDO_REFERENCIA.match(ref_proveedor):
            proveedor_actual = ref_proveedor
            if not actual.proveedor:
                actual.proveedor = ref_proveedor

        if actual is None:
            continue

        if observacion and observacion not in actual.notas:
            actual.notas.append(observacion)

        if not actual.empaque:
            actual.empaque = empaque_de(celdas[11])

        # Costo de compra: se guarda aparte, nunca se expone en el PDF.
        costo_cantidad = entero(celdas[11])
        if costo is not None:
            actual.costos.append(
                {
                    "cantidad": costo_cantidad or 1,
                    "unitario": round(costo, 2),
                    "proveedor": proveedor_actual or actual.proveedor,
                }
            )

        if unitario is None or unitario <= 0:
            continue

        # "CLISE PARA LOGO" aparece dentro del bloque de un precinto pero es un
        # servicio propio: se saca a su propio producto para poder cotizarlo
        # como línea independiente.
        if PATRON_CLISE.search(cantidad_texto) or PATRON_CLISE.search(referencia):
            servicio = rec.obtener("CLISÉ PARA MARCACIÓN DE LOGO", HOJA_PRECIOS)
            servicio.escalones.append(Escalon(cantidad=1, unitario=unitario))
            continue

        cantidad = entero(cantidad_cruda)
        if cantidad is None or cantidad <= 0:
            rec.anotar(
                "cantidad_ilegible",
                f"{actual.nombre}: cantidad '{cantidad_texto}' con precio {unitario}",
                producto=actual.id,
            )
            continue

        if cantidad_texto and not re.fullmatch(r"[\d.,\s]+", cantidad_texto):
            rec.anotar(
                "cantidad_con_texto",
                f"{actual.nombre}: se interpretó '{cantidad_texto}' como {cantidad}",
                producto=actual.id,
            )

        # Coherencia aritmética del Excel: el total sin IVA debería ser
        # unitario × cantidad. Se reporta, no se corrige: manda el unitario.
        subtotal_excel = numero(celdas[4])
        esperado = unitario * cantidad
        if subtotal_excel is not None and abs(subtotal_excel - esperado) > 1:
            rec.anotar(
                "subtotal_inconsistente",
                f"{actual.nombre} @{cantidad}: Excel dice {subtotal_excel:.0f}, "
                f"unitario×cantidad da {esperado:.0f}",
                producto=actual.id,
            )

        actual.escalones.append(
            Escalon(
                cantidad=cantidad,
                unitario=round(unitario, 2),
                logo=logo_actual,
                nota=observacion or None,
            )
        )


# --- Lectura de la hoja de etiquetas --------------------------------------


def leer_etiquetas(hoja, rec: Recolector) -> None:
    """
    Columnas: A referencia · B cantidad · C medida · D unitario · E subtotal ·
    F iva · G total · H centímetro adicional.

    Las etiquetas se cotizan por medida además de por volumen, así que la
    medida viaja dentro del escalón y el producto declara sus medidas.
    """
    actual: Producto | None = None
    medida_actual = ""

    for fila in hoja.iter_rows(min_row=2, values_only=True):
        celdas = list(fila) + [None] * (10 - len(fila))
        referencia = texto(celdas[0])
        medida = texto(celdas[2])
        unitario = numero(celdas[3])
        cm_adicional = numero(celdas[7])
        nota = texto(celdas[9])

        if referencia and not RUIDO_REFERENCIA.match(referencia):
            actual = rec.obtener(referencia, HOJA_ETIQUETAS)
            actual.origen = HOJA_ETIQUETAS
            medida_actual = ""

        if actual is None:
            continue

        if medida:
            medida_actual = medida
            actual.medidas.setdefault(medida, None)

        if nota and nota not in actual.notas:
            actual.notas.append(nota)

        # El recargo por centímetro adicional se declara en la fila donde
        # arranca cada medida y aplica a esa medida.
        if cm_adicional and medida_actual:
            actual.medidas[medida_actual] = cm_adicional

        cantidad = entero(celdas[1])
        if unitario is None or cantidad is None or cantidad <= 0:
            continue

        actual.escalones.append(
            Escalon(
                cantidad=cantidad,
                unitario=round(unitario, 2),
                logo=True,
                medida=medida_actual or None,
            )
        )


# --- Lectura de satelitales ------------------------------------------------


def leer_satelitales(hoja) -> dict:
    """
    Los equipos GPS no se cotizan por unidades sino por número de equipos, con
    planes de plataforma y tarifas de alquiler. Modelo aparte.
    """
    equipos: list[dict] = []
    planes: list[dict] = []
    alquiler: list[dict] = []
    seccion = "venta"

    for fila in hoja.iter_rows(values_only=True):
        celdas = list(fila) + [None] * (4 - len(fila))
        etiqueta = texto(celdas[0])
        if not etiqueta:
            continue

        clave = etiqueta.upper()
        if clave.startswith("VENTA"):
            seccion = "venta"
            continue
        if clave.startswith("ALQUILER"):
            seccion = "alquiler"
            continue

        valores = [numero(c) for c in celdas[1:4]]
        if not any(v for v in valores):
            continue

        if seccion == "alquiler":
            alquiler.append({"periodo": etiqueta, "valor": valores[0]})
        elif clave.startswith("SERVICIO"):
            planes.append(
                {
                    "nombre": etiqueta,
                    "tiers": _tiers_satelital(valores),
                }
            )
        else:
            equipos.append({"modelo": etiqueta, "tiers": _tiers_satelital(valores)})

    return {"equipos": equipos, "planes": planes, "alquiler": alquiler}


def _tiers_satelital(valores: list[float | None]) -> list[dict]:
    rangos = [(1, "1 equipo"), (2, "2 a 9 equipos"), (10, "10 equipos o más")]
    return [
        {"desde": desde, "etiqueta": etiqueta, "valor": valor}
        for (desde, etiqueta), valor in zip(rangos, valores)
        if valor
    ]


# --- Consolidación ---------------------------------------------------------


def consolidar_escalones(producto: Producto, rec: Recolector) -> list[dict]:
    """
    Deja un solo precio por combinación de cantidad, logo y medida.

    Cuando el Excel trae la misma combinación dos veces con precios distintos
    —pasa en varios bloques— se conserva **la última fila**, no la más barata.

    Antes ganaba la más barata, y eso se comporta justo al revés de lo que hace
    falta: si un proveedor sube y en la hoja conviven la fila vieja y la nueva,
    quedarse con la mínima significa cotizar siempre al precio viejo. Al editar
    una hoja de cálculo lo nuevo se escribe debajo, así que la última fila es la
    apuesta razonable. Sigue quedando constancia en las incidencias, porque la
    ambigüedad la resuelve el Excel, no este script.
    """
    mejores: dict[tuple, Escalon] = {}
    for escalon in producto.escalones:
        clave = escalon.clave()
        previo = mejores.get(clave)
        if previo is not None and abs(previo.unitario - escalon.unitario) > 0.01:
            rec.anotar(
                "escalon_duplicado",
                f"{producto.nombre} @{escalon.cantidad}"
                f"{' con logo' if escalon.logo else ''}: "
                f"{previo.unitario:.0f} y {escalon.unitario:.0f}; "
                f"se usa {escalon.unitario:.0f}, que es el de la fila de más abajo",
                producto=producto.id,
            )
        mejores[clave] = escalon

    ordenados = sorted(
        mejores.values(), key=lambda e: (e.medida or "", bool(e.logo), e.cantidad)
    )

    # Comprar más debería salir más barato por unidad. Donde no ocurre, casi
    # siempre es un dedazo en el Excel, y conviene que alguien lo mire antes de
    # que salga en una cotización.
    for anterior, siguiente in zip(ordenados, ordenados[1:]):
        mismo_bloque = (anterior.medida or "") == (siguiente.medida or "") and bool(
            anterior.logo
        ) == bool(siguiente.logo)
        if mismo_bloque and siguiente.unitario > anterior.unitario:
            rec.anotar(
                "precio_no_monotono",
                f"{producto.nombre}: {anterior.cantidad} uds a "
                f"{anterior.unitario:.0f} pero {siguiente.cantidad} uds a "
                f"{siguiente.unitario:.0f}",
                producto=producto.id,
            )

    return [
        {
            "desde": e.cantidad,
            "unitario": e.unitario,
            **({"logo": e.logo} if e.logo is not None else {}),
            **({"medida": e.medida} if e.medida else {}),
            **({"nota": e.nota} if e.nota else {}),
        }
        for e in ordenados
    ]


def version_de(productos: list[dict]) -> str:
    """Huella corta de los precios publicados."""
    relevante = [
        {"id": p["id"], "escalones": p["escalones"]}
        for p in sorted(productos, key=lambda p: p["id"])
    ]
    serializado = json.dumps(relevante, ensure_ascii=False, sort_keys=True)
    return hashlib.sha256(serializado.encode("utf-8")).hexdigest()[:12]


def costo_minimo(producto: Producto) -> float | None:
    valores = [c["unitario"] for c in producto.costos if c["unitario"] > 0]
    return min(valores) if valores else None


def construir(ruta: Path, incluir_costos: bool = True) -> dict:
    libro = openpyxl.load_workbook(ruta, data_only=True)
    rec = Recolector()
    comerciales = cargar_nombres_comerciales()

    leer_listado_precios(libro[HOJA_PRECIOS], rec)
    if HOJA_ETIQUETAS in libro.sheetnames:
        leer_etiquetas(libro[HOJA_ETIQUETAS], rec)

    productos = []
    for producto in rec.productos.values():
        escalones = consolidar_escalones(producto, rec)
        if not escalones:
            rec.anotar("producto_sin_precio", producto.nombre, producto=producto.id)
            continue

        con_logo = [e for e in escalones if e.get("logo")]
        sin_logo = [e for e in escalones if e.get("logo") is False]

        # El nombre que ve el cliente sale del catálogo comercial; el listado
        # de precios pone el id, la categoría y los escalones. Cuando no hay
        # nombre comercial, la referencia hace las dos veces y `referencia` se
        # omite: repetir el mismo texto en dos campos no informa de nada.
        comercial = comerciales.get(producto.id)
        productos.append(
            {
                "id": producto.id,
                "nombre": comercial["nombre"] if comercial else producto.nombre,
                **({"referencia": producto.nombre} if comercial else {}),
                "categoria": producto.categoria,
                "escalones": escalones,
                "minimo": min(e["desde"] for e in escalones),
                "admiteLogo": bool(con_logo),
                "admiteSinLogo": bool(sin_logo) or not con_logo,
                **(
                    {
                        "medidas": [
                            {
                                "nombre": nombre,
                                **({"cmAdicional": cm} if cm else {}),
                            }
                            for nombre, cm in producto.medidas.items()
                        ]
                    }
                    if producto.medidas
                    else {}
                ),
                **({"empaque": producto.empaque} if producto.empaque else {}),
                **({"proveedor": producto.proveedor} if incluir_costos and producto.proveedor else {}),
                **(
                    {"costoReferencia": costo_minimo(producto)}
                    if incluir_costos and costo_minimo(producto)
                    else {}
                ),
                **({"notas": producto.notas} if producto.notas else {}),
            }
        )

    productos.sort(key=lambda p: (peso_categoria(p["categoria"]), p["nombre"]))
    categorias = sorted({p["categoria"] for p in productos}, key=peso_categoria)

    return {
        "generadoEl": datetime.now().date().isoformat(),
        # Huella de los precios, no de la fecha: regenerar el catálogo sin que
        # nada haya cambiado deja la misma versión. La aplicación la usa para
        # saber si una cotización guardada se armó con otra lista de precios.
        "version": version_de(productos),
        "origen": ruta.name,
        "iva": IVA,
        "categorias": categorias,
        "productos": productos,
        "satelitales": (
            leer_satelitales(libro[HOJA_SATELITALES])
            if HOJA_SATELITALES in libro.sheetnames
            else {"equipos": [], "planes": [], "alquiler": []}
        ),
        "incidencias": rec.incidencias,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("excel", type=Path, help="Listado de precios .xlsx")
    parser.add_argument(
        "-o",
        "--salida",
        type=Path,
        default=Path("src/datos/catalogo.json"),
        help="Ruta del JSON de salida",
    )
    parser.add_argument(
        "--sin-costos",
        action="store_true",
        help=(
            "Omite costo de compra y proveedor. Úsalo si el catálogo va a salir "
            "de la máquina del área comercial."
        ),
    )
    args = parser.parse_args()

    catalogo = construir(args.excel, incluir_costos=not args.sin_costos)
    args.salida.parent.mkdir(parents=True, exist_ok=True)
    args.salida.write_text(
        json.dumps(catalogo, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    escalones = sum(len(p["escalones"]) for p in catalogo["productos"])
    print(f"{len(catalogo['productos'])} productos · {escalones} escalones")
    print(f"{len(catalogo['categorias'])} categorías: {', '.join(catalogo['categorias'])}")
    sin_comercial = [p["nombre"] for p in catalogo["productos"] if "referencia" not in p]
    if sin_comercial:
        print(f"{len(sin_comercial)} referencias sin nombre comercial (salen como en el listado):")
        for nombre in sin_comercial:
            print(f"  · {nombre}")
    print(f"{len(catalogo['incidencias'])} incidencias registradas")
    print(f"→ {args.salida}")


if __name__ == "__main__":
    main()
